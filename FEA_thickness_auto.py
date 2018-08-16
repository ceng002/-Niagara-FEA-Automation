# FEA_thickness_auto.PY Read Me
# last edit 2018-7-9
# Compiled for:
# Python        3.6
# numpy		    1.14.5
# matplotlib	2.2.2
# openpyxl	    2.5.4
#
# Instructions:
# 1.If possible, place the Excel or csv files titled "loc_x", "loc_y", "loc_z" and "raw_thickness" into the parent
#   folder of this program(the file in which this program is stored).
# 2.Open the python file FEA_thickness_Auto.py. Ensure that openpyxl, numpy, and matplotlib
#   are all installed on the IDE.
# 3.Run the program. If it can't find the files, the user will need to specify them manually below as strings.
#   Please use file extensions! Acceptable extensions are .csv and .xlsx. If the files are in another location,
#   the full file path should be used. Also, please remember to keep the r-prefix on the string.
# 4.If the program is successful, it should create a file in the parent folder.
# 5.Open the file and verify that the information is correct (Height should increment in multiples of 5, and
#   repeat 72 times. Angle should count from 0 to 355 in increments of 5. Radius and Thickness should
#   be reasonable in mm.

import openpyxl as xl
import csv
import numpy as np
import matplotlib.pyplot as plt

# user parameters:
# file paths can be manually specified here (keep lowercase r in front of string!)
locX_fp = r"loc_x.xlsx"
locY_fp = r"loc_y.xlsx"
locZ_fp = r"loc_z.xlsx"
th_fp = r"raw_thickness.xlsx"

# upper and lower trim values for base and finish (multiples of 5).
min_height = 15
max_height = 180

# title of generated csv file (include .csv file extension)
outTitle = 'final_radial_thickness2.csv'
# stops file generation and instead prints file to console
debug = False
# END user parameters

# BEGIN program:
# force numpy to use decimal instead of scientific notation in console
np.set_printoptions(suppress=True)
# lists to simplify console prints
debugFile = ['locX', 'locY', 'locZ', 'thickness']
debugHLD = ['Height', 'Length', 'Depth']
debugHeader = ['Height', 'Angle', 'Radius', 'Thickness']
# read files:
filePath = [locX_fp, locY_fp, locZ_fp, th_fp]

# create  and initialize a list of data handles and metadata:
fileHandle = []
fileType = []
fileLen = []

for i in range(0, len(filePath)):
    fileHandle.append(None)
    fileType.append(None)
    fileLen.append(None)

print('attempting to read files:')

for i in range(0, len(filePath)):
    if filePath[i].find('.csv') > 0:
        fileHandle[i] = (csv.reader(open(filePath[i])))
        fileType[i] = 'csv'
        fileLen[i] = sum(1 for row in csv.reader(open(filePath[i])))
        print('loaded ' + debugFile[i] + ' as csv.')

    elif filePath[i].find('.x') > 0:
        fileHandle[i] = xl.load_workbook(filePath[i]).active
        fileType[i] = 'xl'
        fileLen[i] = fileHandle[i].max_row
        print('loaded ' + debugFile[i] + ' as excel.')
    else:
        print(debugFile[i] + ' not loaded.')

# sub 1 for header in x, y, z.
for i in range(0, 3):
    fileLen[i] -= 1

# create an array of zeroes to populate for thickness
thickness = np.zeros([fileLen[3], 9], dtype='float')

# load thickness data
if fileType[3] == 'xl':
    for i in range(0, fileLen[3]):
        for j in range(0, 9):
            thickness[i, j] = fileHandle[3].cell(row=i+1, column=j+3).value
    #print('thickness:\n' + str(thickness))
elif fileType[3] == 'csv':
    i = 0
    for row in fileHandle[3]:
        j = 0
        for th in row:
            thickness[i, j] = th
            #print(str(th) + ",")
            j += 1
        i += 1

# length check
for i in range(0, 4):
    print('Length of ' + debugFile[i] + ': ' + str(fileLen[i]))
# error catch
if not fileLen[0] == fileLen[1] == fileLen[2]:
    print('Lengths of hld data are not equal, halting.')
    exit()

# read xyz data
lengthXYZ = fileLen[0]
print('lengthXYZ: ' + str(lengthXYZ))

arrXYZ = []

for i in range(0, 3):
    arrXYZ.append(np.zeros([fileLen[0], 2]))

    if fileType[i] == 'xl':
        for j in range(0, lengthXYZ):
            arrXYZ[i][j, 0] = fileHandle[i].cell(row=j+2, column=1).value
            arrXYZ[i][j, 1] = fileHandle[i].cell(row=j+2, column=2).value

    elif fileType[i] == 'csv':
        j = -2
        for row in fileHandle[i]:
            j += 1
            if j == -1:
                continue
            arrXYZ[i][j, 0] = row[0]
            arrXYZ[i][j, 1] = row[1]

# sort by node number
for arr in arrXYZ:
    arr = np.asarray(sorted(arr, key=lambda p: p[0]))

# check that all nodes match
for i in range(0, lengthXYZ):
    if not arrXYZ[0][i, 0] == arrXYZ[1][i, 0] == arrXYZ[2][i, 0]:
        print('nodes do not align, halting.')
        print('index:\t ' + str(i))
        for j in range(0, 3):
            print(debugFile[j] + '\t' + str(int(arrXYZ[j][i, 0])))
        exit()

# generate height, radius, depth array
hld = np.zeros([lengthXYZ, 3])
print('hld length: ' + str(len(hld)))

# sort by range
rangeXYZ = []
# find ranges
for i in range(0, 3):
    val = arrXYZ[i][:, 1]
    rangeXYZ.append(np.max(val) - np.min(val))

orderXYZ = np.argsort(rangeXYZ)[::-1]
for i in range(0, len(orderXYZ)):
    print('using ' + debugFile[i] + ' as ' + debugHLD[orderXYZ[i]] + ': range ' + str(rangeXYZ[i]))

# height has the largest range, then length, then depth
#print('\nhld:\n' + str(rangeXYZ))
for i in range(0, 3):
    for j in range(0, len(hld)):
        hld[j, i] = arrXYZ[orderXYZ[i]][j, 1]
#print(hld)

# sort along height
hld = np.asarray(sorted(hld, key=lambda p: p[0]))
#print(hld)

# trim base and finish values
print('\nradius fit lower bound: ' + str(min_height))
print('radius fit upper bound: ' + str(max_height))

index_base = 0
index_finish = -1

while hld[index_base, 0] < min_height:
    index_base += 1

while hld[index_finish, 0] > max_height:
    index_finish -= 1

if index_finish != -1:
    index_finish += 1
    print('\ntrimming above upper height: ' + str(hld[index_finish, 0]) + ' at index ' + str(index_finish + len(hld)))
    hld = np.delete(hld, np.s_[index_finish + len(hld) + 1:len(hld)], axis=0)

if index_base != 0:
    index_base -= 1
    print('trimming below lower height: ' + str(hld[index_base, 0]) + ' at index ' + str(index_base))
    hld = np.delete(hld, np.s_[0:index_base], axis=0)

print('trimmed length: ' + str(hld.shape[0]))
#print(hld)

# generate arrays for radius macro
# r =  sqrt(l^2 + d^2)
r = np.zeros(len(hld))
for i in range(0, len(hld)):
    r[i] = (hld[i, 1] ** 2 + hld[i, 2] ** 2) ** 0.5

# BEGIN radius_macro_1GAL.PY
# used later in final_radial_thickness.csv
rowCount = int((max_height - min_height)/5) + 1

x = np.linspace(max_height, min_height, rowCount, endpoint=True, dtype='int_')
y = np.zeros([rowCount])

j = len(hld) - 1
for i in range(0, len(x)):
    while x[i] < hld[j, 0]:
        j -= 1
        #print(h[j])
    m = (r[j+1] - r[j]) / (hld[j+1, 0] - hld[j, 0])
    dh = x[i] - hld[j, 0]
    y[i] = r[j] + (m * dh)
#    print("h index: " + str(j))
#    print('matching ' + str(x[i]) + ' with ' + str(hld[j, 2]))
#    print('h: '+ str(h[j]))
#    print('m: ' + str(m))
#    print('dh: ' + str(dh))
#    print('dr:' + str(m*dh))

print('radius samples: ' + str(len(x)))

#print('h values\n' + str(x))
#print('r values:\n' + str(y))
plt.plot(hld[:, 0], r, '.', x, y, 'o')
plt.show()

radius_full = np.zeros([len(x), 72])

for i in range(0, len(x)):
    for j in range(0, 72):
        radius_full[i, j] = y[i]

# END radius_macro_1GAL.PY

# generate arrays for thickness_macro_1GAL.PY

#print('t:\n' + str(thickness))
# BEGIN thickness_macro_1GAL.PY

angle = np.array([0, 45, 90, 135, 180, 225, 270, 315, 360])
# generate final sub-arrays
# rowCount from thickness macro
thickness_full = np.zeros((rowCount, 72))
height_full = np.zeros((rowCount, 72))
angle_full = np.zeros((rowCount, 72))

for i in range(0, np.size(x)):
    for j in range(0, np.size(angle) - 1):
        x1 = angle[j]
        x2 = angle[j + 1]
        y1 = thickness[i, j]
        y2 = thickness[i, j + 1]
        k = (y1 - y2) / (x1 - x2)
        b = y1 - k * x1
        angle_temp = np.linspace(x1, x2, 10)
        for z in range(0, np.size(angle_temp)-1):
            x3 = angle_temp[z]
            y3 = k * x3 + b
            thickness_full[i, j * 9 + z] = y3
            height_full[i, j * 9 + z] = x[i]
            angle_full[i, j * 9 + z] = x3
# END thickness_macro_1GAL.PY

print('\nfinal checks:')

print('height length:\t\t' + str(height_full.shape[0]))
print('angle length:\t\t' + str(angle_full.shape[0]))
print('radius length:\t\t' + str(radius_full.shape[0]))
print('thickness length:\t' + str(thickness_full.shape[0]))

print('\nheight width:\t\t' + str(height_full.shape[1]))
print('angle width:\t\t' + str(angle_full.shape[1]))
print('radius width:\t\t' + str(radius_full.shape[1]))
print('thickness width:\t' + str(thickness_full.shape[1]))

# generate final array
out = np.zeros([rowCount*72, 4], dtype='float_')
for i in range(0, rowCount):
    for j in range(0, 72):
        k = i*72 + j
        out[k, 0] = height_full[i, j]
        out[k, 1] = angle_full[i, j]
        out[k, 2] = radius_full[i, j]
        out[k, 3] = thickness_full[i,j]

head = ''
for i in range(0, len(debugHeader)):
    head += debugHeader[i] + ','
head = head[0:len(head) - 1]

if debug:
    print('\n' + outTitle + ":")
    print(head)
    print(str(out))
    print('\nno files generated because debug = True.')
else:
    np.savetxt(outTitle, out, comments="", fmt='%f', delimiter=',', header=head)
    print("\nprocess completed. Find generated file in the parent file of this program with title:\n\t" + outTitle)