# FEA_fileGenerator.PY ReadMe
# last edit 2018-7-5
# Compiled for:
# Python        3.6
# numpy		    1.14.5
# matplotlib	2.2.2
# openpyxl	    2.5.4
#
# This file is intended for use in 'SOP: FEM Feel-Stiffness Analysis' compiled by Aparna Avalli.

import openpyxl as xl
import numpy as np
import matplotlib.pyplot as plt
import numpy.polynomial.polynomial as poly

# user parameters:
# set true if reading from Excel files
xlFlag = True
# excel file paths only
locX_filePath = r"C:\Users\cheng\Desktop\Automation FEA-Feel Stiffness\loc_x.xlsx"
locY_filePath = r"C:\Users\cheng\Desktop\Automation FEA-Feel Stiffness\loc_y.xlsx"
locZ_filePath = r"C:\Users\cheng\Desktop\Automation FEA-Feel Stiffness\loc_z.xlsx"
thickness_filePath = r"C:\Users\cheng\Desktop\Automation FEA-Feel Stiffness\raw_thickness.xlsx"
# upper and lower trim values for base and finish
min_height = 15
max_height = 180
# title of generated file (include .csv file extension)
outTitle = 'final_radial_thickness1.csv'
# header in generated file (leave blank to skip)
head = "Height,Angle,Radius,Thickness"

# stops file generation and instead prints file to console
debug = False

# BEGIN program:
# force np to use decimal instead of scientific notation in console
np.set_printoptions(suppress=True)

# read files:
lengthX = lengthY = lengthZ = lengthT = 0
print('attempting to read local csv files:')
csvError = 'locX'
try:
    csv_locX = open('locX.csv')
    csvError = 'locY'
    csv_locY = open('locY.csv')
    csvError = 'locY'
    csv_locZ = open('locZ.csv')
    csvError = 'locY'
    csv_thickness = open('raw_thickness.csv')
    print("loaded all csv files.")

    lengthX = sum(1 for row in csv_locX)
    lengthY = sum(1 for row in csv_locY)
    lengthZ = sum(1 for row in csv_locZ)
    lengthT = sum(1 for row in csv_thickness)
except:
    print('load ' + csvError + ' failed, attempting to read Excel files:')

xlError = "locX"
sheetX = sheetY = sheetZ = sheetT = None
try:
    locX = xl.load_workbook(locX_filePath)
    sheetX = locX.active

    xlError = "locY"
    locY = xl.load_workbook(locY_filePath)
    sheetY = locY.active

    xlError = "locZ"
    locZ = xl.load_workbook(locZ_filePath)
    sheetZ = locZ.active

    xlError = "thickness"
    thickness = xl.load_workbook(thickness_filePath)
    sheetT = thickness.active
    print('loaded all Excel files.')
    xlFlag = True

    # sub 1 for header
    lengthX = sheetX.max_row - 1
    lengthY = sheetY.max_row - 1
    lengthZ = sheetZ.max_row - 1
    lengthT = sheetT.max_row - 1
except:
    print(xlError + ' file not found, halting.')
    exit()


print('\nlength of locX: ' + str(lengthX))
print('length of locY: ' + str(lengthY))
print('length of locZ: ' + str(lengthZ))
print('length of thickness: ' + str(lengthT))
# error catch
if not lengthX == lengthY == lengthZ:
    print('Lengths are not equal, halting.')
    exit()

# aggregate xyz positions

lengthXYZ = lengthX

arrX = np.zeros([lengthXYZ, 2])
arrY = np.zeros([lengthXYZ, 2])
arrZ = np.zeros([lengthXYZ, 2])
if xlFlag:
    for i in range(0, len(arrX)):
        arrX[i, 0] = sheetX.cell(row=i+2, column=1).value
        arrX[i, 1] = sheetX.cell(row=i+2, column=2).value

    for i in range(0, len(arrY)):
        arrY[i, 0] = sheetY.cell(row=i+2, column=1).value
        arrY[i, 1] = sheetY.cell(row=i+2, column=2).value

    for i in range(0, len(arrZ)):
        arrZ[i, 0] = sheetZ.cell(row=i+2, column=1).value
        arrZ[i, 1] = sheetZ.cell(row=i+2, column=2).value

# sort by node number
arrX = np.asarray(sorted(arrX, key=lambda p: p[0]))
arrY = np.asarray(sorted(arrY, key=lambda p: p[0]))
arrZ = np.asarray(sorted(arrZ, key=lambda p: p[0]))

# check that all nodes match
for i in range(0, lengthXYZ):
    if not arrX[i, 0] == arrY[i, 0] == arrZ[i, 0]:
        print('nodes do not align, halting')
        exit()

# generate position array
xyz = np.zeros([lengthX, 3])
print('xyz length: ' + str(len(xyz)))


pos = [arrX, arrY, arrZ]
for i in range(0, len(xyz)):
    xyz[i, 0] = arrX[i, 1]
    xyz[i, 1] = arrY[i, 1]
    xyz[i, 2] = arrZ[i, 1]

# sort along Z
xyz = np.asarray(sorted(xyz, key=lambda p: p[2]))

# trim base and finish values
index_base = 0
index_finish = -1

while xyz[index_base, 2] < min_height:
    index_base += 1
print('\ntrimming below lower height: ' + str(xyz[index_base, 2]) + ' at index ' + str(index_base))

while xyz[index_finish, 2] > max_height:
    index_finish -= 1
print('trimming above upper height: ' + str(xyz[index_finish, 2]) + ' at index ' + str(index_finish + xyz.shape[0]))

xyz = np.delete(xyz, np.s_[index_finish + xyz.shape[0]:xyz.shape[0]], axis=0)
xyz = np.delete(xyz, np.s_[0:index_base], axis=0)
print('trimmed length: ' + str(xyz.shape[0]))
# print(xyz)

# generate arrays for radius macro
# h = z-axis
h = np.zeros(len(xyz))
for i in range(0, len(xyz)):
    h[i] = xyz[i,2]
# r =  sqrt(x^2 + y^2)
r = np.zeros(len(xyz))
for i in range(0, len(xyz)):
    r[i] = (xyz[i,0] ** 2 + xyz[i,1] ** 2) ** 0.5

# BEGIN radius_macro_1GAL.PY
# find bounds for x linspace (should be multiples of 5 inclusive of all points in h)
l = h[0].astype('int_')
l = l - (l % 5)
u = h[-1].astype('int_')
u = u - (u % 5) + 5

# used later in final_radial_thickness.csv
rowCount = int((u-l)/5) + 1

print('\nradius fit lower bound: ' + str(l))
print('radius fit upper bound: ' + str(u))


x = np.linspace(u, l, rowCount, endpoint=True, dtype='int_')
y = poly.polyval(x, poly.polyfit(h, r, 100))
print('radius samples: ' + str(len(x)))
#print('x:\n' + str(x))
#print('y:\n' + str(y))
#print(x)

plt.plot(h, r, '.', x, y, '-')
plt.show()

R = np.zeros([len(x), 72])

for i in range(0, len(x)):
    for j in range(0, 72):
        R[i, j] = y[i]

# END radius_macro_1GAL.PY

# generate arrays for thickness_macro_1GAL.PY
thickness = np.zeros([lengthT, 9], dtype='float')

for i in range(0, lengthT):
    for j in range(0, 9):
        thickness[i, j] = sheetT.cell(row=i+1, column=j+1).value
#print('thickness:\n' + str(thickness))

# BEGIN thickness_macro_1GAL.PY

angle = np.array([0, 45, 90, 135, 180, 225, 270, 315, 360])
# generate final sub-arrays
# rowCount from thickness macro
thickness_full = np.zeros((rowCount, 72))
height_full = np.zeros((rowCount, 72))
angle_full = np.zeros((rowCount, 72))

for i in range(0, np.size(x)):
    for j in range(0, np.size(angle) - 1):
        x1 = angle[j]
        x2 = angle[j + 1]
        y1 = thickness[i, j]
        y2 = thickness[i, j + 1]
        k = (y1 - y2) / (x1 - x2)
        b = y1 - k * x1
        angle_temp = np.linspace(x1, x2, 10)
#        print(angle_temp)
        for z in range(0, np.size(angle_temp)-1):
            x3 = angle_temp[z]
            y3 = k * x3 + b
            thickness_full[i, j * 9 + z] = y3
#            print(y3)
            height_full[i, j * 9 + z] = x[i]
            angle_full[i, j * 9 + z] = x3
# BEGIN thickness_macro_1GAL.PY


print('\nfinal checks:')
print('height length:\t\t' + str(height_full.shape[0]))
print('angle length:\t\t' + str(angle_full.shape[0]))
print('radius length:\t\t' + str(R.shape[0]))
print('thickness length:\t' + str(thickness_full.shape[0]))

print('\nheight width:\t\t' + str(height_full.shape[1]))
print('angle width:\t\t' + str(angle_full.shape[1]))
print('radius width:\t\t' + str(R.shape[1]))
print('thickness width:\t' + str(thickness_full.shape[1]))

# generate final array
out = np.zeros([rowCount*72, 4], dtype='float_')
for i in range(0, rowCount):
    for j in range(0, 72):
        k = i*72 + j
        out[k, 0] = height_full[i, j]
        out[k, 1] = angle_full[i, j]
        out[k, 2] = R[i, j]
        out[k, 3] = thickness_full[i, j]

if debug:
    print('\n' + outTitle + ":")
    if head:
        print(head)
    print(str(out))
    print('\nno files generated because debug = True.')
else:
    np.savetxt(outTitle, out, comments="", fmt='%f', delimiter=',', header=head)
    print("\nprocess completed. Find generated file in the parent file of this program with title:\n\t" + outTitle)